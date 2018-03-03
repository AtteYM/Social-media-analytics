from openpyxl import load_workbook
import collections
import xlsxwriter

#Goes through string and makes a list containing all tags which start with '#'
def get_hashtagslist(string):
    ret = []
    s=''
    hashtag = False
    for char in string:
        if char=='#':
            hashtag = True
            if s:
                ret.append(s)
                s=''
 #           continue
 #remove previous commentmark if you want to exclude # from words

        #define characters that aren´t included in tags
        #in other words which characters define end of tag
        if hashtag and char in [' ',"'",'’','!','|','\n','.',',','(',')',':','{','}, '] and s:
            ret.append(s)
            s=''
            hashtag=False
        #save tag in to list
        if hashtag:
            s+=char

    #save last tag in to list
    if s:
        ret.append(s)

    #return list of tags used in tweet
    return list(set([word for word in ret if len(ret)>1 and len(ret)<20]))

def main():

    #open specific workbook and sheet
    wb = load_workbook('Tampere3_Luuppi.xlsx')
    ws = wb['DataTaulu']

    htags = []

    #save column and go through specific rows containing tweets
    colC = ws['C']
    for i in range(1,1728):
        #for cleaning the data make words lowercase
        tweet = str(colC[i].value).lower()
        htags = htags + get_hashtagslist(tweet)

    #turn list in to counter to group tags and calculate frequency
    counter = collections.Counter(htags)

    #print most common tags and amount of unique tags
    #print(counter.most_common(10))
    #print(len(counter))

    #write tags and frequencies in to first two colums of a new file
    workbook = xlsxwriter.Workbook('data2.xlsx')
    worksheet = workbook.add_worksheet()

    row = 0
    col = 0

    for key in counter.keys():
        row += 1
        worksheet.write(row, col, key)
        worksheet.write(row, col + 1, counter[key])

    workbook.close()

main()
